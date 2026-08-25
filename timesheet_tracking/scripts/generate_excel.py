"""Create June, July, and August 2026 sample workbooks for Timewise."""
from calendar import monthrange
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FIRST = ['Aarav','Aditi','Akash','Ananya','Arjun','Bhavya','Charan','Deepa','Dev','Divya','Esha','Farhan','Gautam','Harini','Ishaan','Jaya','Karthik','Kavya','Krishna','Lakshmi','Manoj','Meera','Naveen','Nisha','Pooja','Pranav','Priya','Rahul','Rhea','Rohan','Sahana','Sanjay','Shalini','Siddharth','Sneha','Sonia','Suraj','Tanvi','Tejas','Uma','Varun','Vidya','Vikram','Yamini','Yash','Zara','Anil','Bala','Chitra','Dinesh']
LAST = ['Sharma','Iyer','Patel','Reddy','Kumar','Nair','Rao','Menon','Das','Singh']
PROJECTS = [('PRJ-101','Atlas CRM','Customer Platforms'),('PRJ-102','Nova Mobile','Digital Experience'),('PRJ-103','Orion Data Hub','Data & Analytics'),('PRJ-104','Helios Cloud','Infrastructure'),('PRJ-105','Pulse Payments','Finance Technology'),('PRJ-106','Aster Support','Operations'),('PRJ-107','Lumen AI','Innovation Lab'),('PRJ-108','Harbor ERP','Enterprise Systems'),('PRJ-109','Vertex Security','Cybersecurity'),('PRJ-110','Echo Commerce','Digital Experience')]
PLANS = {
    6: ({1:{18:'Annual leave',19:'Annual leave'},8:{4:'Sick leave'},14:{22:'Personal leave'},22:{10:'Annual leave',11:'Annual leave'},35:{29:'Sick leave'},43:{15:'Annual leave'}},{3:[12],9:[24,25],17:[5],28:[16,17],39:[30],46:[8]}),
    7: ({4:{9:'Annual leave',10:'Annual leave'},11:{20:'Sick leave'},18:{27:'Annual leave'},24:{15:'Personal leave'},33:{6:'Sick leave'},41:{28:'Annual leave',29:'Annual leave'}},{0:[17],7:[1,2],15:[24],21:[13,14],30:[31],48:[21]}),
    8: ({2:{11:'Annual leave',12:'Annual leave'},7:{21:'Sick leave'},13:{17:'Annual leave',18:'Annual leave',19:'Annual leave'},19:{28:'Personal leave'},26:{6:'Sick leave'},31:{24:'Annual leave'},42:{13:'Personal leave',14:'Personal leave'}},{4:[7,20],10:[14],16:[24,25],23:[5],29:[18,19],37:[28],47:[11,12]})
}

def style(sheet):
    for cell in sheet[1]:
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='173B53'); cell.alignment=Alignment(horizontal='center')
    sheet.freeze_panes='A2'; sheet.auto_filter.ref=sheet.dimensions
    for col in range(1,sheet.max_column+1): sheet.column_dimensions[get_column_letter(col)].width=18

def build(month):
    leaves, missing = PLANS[month]; dates=[date(2026,month,day) for day in range(1,monthrange(2026,month)[1]+1) if date(2026,month,day).weekday()<5]
    def week(d): return f'W{dates.index(d)//5+1}'
    wb=Workbook(); people=wb.active; people.title='Employees'
    people.append(['Employee ID','Employee Name','Role','Manager','Project 1','Project 2','Expected Work Hours','Booked Hours','Leave Hours','Missing Hours','Status'])
    booking=wb.create_sheet('Daily Bookings'); booking.append(['Date','Week','Employee ID','Employee Name','Project Code','Project','Hours','Entry Type'])
    leave_sheet=wb.create_sheet('Leave & Holidays'); leave_sheet.append(['Date','Employee ID','Employee Name','Leave Type','Hours','Notes'])
    project_sheet=wb.create_sheet('Projects'); project_sheet.append(['Project Code','Project','Portfolio'])
    summary=wb.create_sheet('Weekly Summary'); summary.append(['Week','Employee ID','Employee Name','Expected Work Hours','Booked Hours','Leave Hours','Missing Hours','Status'])
    for row in PROJECTS: project_sheet.append(row)
    for i,name in enumerate(FIRST):
        eid=f'EMP-{i+1:03d}'; full=f'{name} {LAST[i%10]}'; p1,p2=PROJECTS[i%10],PROJECTS[(i+3)%10]; approved=leaves.get(i,{}); missed=missing.get(i,[]); booked=leave_hours=0; weekly={}
        for day_index,current in enumerate(dates):
            w=week(current); weekly.setdefault(w,[0,0]);
            if current.day in approved:
                leave_hours+=8; weekly[w][1]+=8; leave_sheet.append([current,eid,full,approved[current.day],8,'Approved sample leave']); continue
            if current.day in missed: continue
            primary=4+(i+day_index)%3
            for project,hours in ((p1,primary),(p2,8-primary)):
                booking.append([current,w,eid,full,project[0],project[1],hours,'Booked']); booked+=hours; weekly[w][0]+=hours
        expected=len(dates)*8-leave_hours; outstanding=expected-booked; status='Missing entries' if outstanding else 'Leave matched' if leave_hours else 'Complete'
        people.append([eid,full,['Software Engineer','Business Analyst','Data Engineer','QA Engineer','Product Designer'][i%5],'Meera Shah',p1[1],p2[1],expected,booked,leave_hours,outstanding,status])
        for w,(hours,leave) in weekly.items():
            planned=len([d for d in dates if week(d)==w])*8-leave; summary.append([w,eid,full,planned,hours,leave,planned-hours,'Missing entries' if planned-hours else 'Leave matched' if leave else 'Complete'])
    for sheet in wb.worksheets: style(sheet)
    output=Path(__file__).resolve().parents[1]/'data'/f'timesheet_{date(2026,month,1):%B_%Y}'.lower()
    output=output.with_suffix('.xlsx'); wb.save(output); print(f'Created {output.name}')

if __name__=='__main__':
    for month in (6,7,8): build(month)
